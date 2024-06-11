from tkinter_utils import print_on_gui
import os
import shutil
import zipfile
import pandas as pd
from datetime import datetime
import time
import xml.etree.ElementTree as Et
import glob
import openpyxl
from pathlib import Path
from move_archives import process_files
import urllib3
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium_utils import SeleniumUtils
from webdriver_manager import WebDriverManager
from pandas_utils import PandasUtils

pu = PandasUtils()


def get_beneficiario_initials_and_numero_lote(xml_file):

    # Register the namespaces to properly parse the XML
    namespaces = {
        'ans': 'http://www.ans.gov.br/padroes/tiss/schemas'
    }

    # Parse the XML file
    tree = Et.parse(xml_file)
    root = tree.getroot()

    # Find the 'nome do beneficiário' element
    nome_beneficiario_element = root.find(".//ans:dadosBeneficiario/ans:nomeBeneficiario", namespaces=namespaces)
    # Find the 'numeroLote' element
    numero_lote_element = root.find(".//ans:loteGuias/ans:numeroLote", namespaces=namespaces)

    # Extract initials from the 'nome do beneficiário'
    if nome_beneficiario_element is not None:
        nome = nome_beneficiario_element.text
        initials = ''.join([part[0].upper() for part in nome.split()])
    else:
        nome = 'Name not found'
        initials = "Initials not found"

    # Extract 'numeroLote'
    if numero_lote_element is not None:
        numero_lote = numero_lote_element.text
    else:
        numero_lote = "NumeroLote not found"

    file_name = f'{initials} - {numero_lote}'
    return file_name, nome, numero_lote


class UploadXML:

    def __init__(self, username='hurbanav', password='Amarelinha12*'):
        driver_manager = WebDriverManager()
        self.driver_path = driver_manager.driver_path
        self.username = username
        self.password = password
        self.orizon_url = "https://sso-fature.orizon.com.br/auth/realms/orizon-dativa/protocol/openid-connect/auth?client_id=fature_client&response_type=code&scope=openid&redirect_uri=https://sso-auth-codeflow-fature-apicast-production.api.cppr.orizon.com.br/sso/token?user_key=32efd36b405a07b8c0e6c6cb9c582047"

    @staticmethod
    def filter_df(df, column_name, column_value, equal=True):
        if equal:
            return df[df[column_name] == column_value]
        else:
            return df[df[column_name] != column_value]

    @staticmethod
    def extract_last_number(text):
        import re
        # Find all occurrences of numbers in the text
        numbers = re.findall(r'\d+', text)
        # Return the last number found, or None if no numbers were found
        return str(numbers[-1]) if numbers else str(5)

    def change_input_value(self,
                           element_locator: tuple,
                           second_element_locator: tuple = None,
                           text=None,
                           waiting_time=10,
                           ):

        if second_element_locator:
            element_to_extract_value = self.su.find_element(*second_element_locator).text
            text = self.extract_last_number(element_to_extract_value)

        input_element = self.su.find_element(*element_locator)
        if input_element is None:
            raise Exception('NoneElement: Clear input value function with invalid element')
        input_element.clear()
        input_element.send_keys(text)
        time.sleep(waiting_time)

    @staticmethod
    def generate_filename():
        # Format today's date as 'YYYY-MM-DD'
        date_str = datetime.today().strftime('%Y-%m-%d')
        filename = f"Upload RPA Run - {date_str}.xlsx"

        # Find the user's home directory
        home = Path.home()

        # Specify the path to the Downloads folder
        # This works for Windows, macOS, and Linux
        downloads_path = home / 'Downloads'

        # Combine the Downloads path with the filename
        filepath = downloads_path / filename

        return str(filepath)

    @staticmethod
    def append_dict_to_excel(filename, data_dict):
        from openpyxl import Workbook
        """
        Appends a dictionary as a new row to an Excel file. If the file does not exist, it creates a new one.
        
        :param filename: Path to the Excel file (should be an .xlsx file)
        :param data_dict: Dictionary containing the data to append. Keys are used as headers.
        """
        # Check if the Excel file exists
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            # If it's the first row, add headers
            if sheet.max_row == 1 and all(cell.value is None for cell in next(sheet.iter_rows())):
                sheet.append(list(data_dict.keys()))
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Adding headers for the new file
            sheet.append(list(data_dict.keys()))

        # Append the data values to the next row
        sheet.append(list(data_dict.values()))

        # Save the workbook
        workbook.save(filename)

    def duplicate_folder(self, original_folder_path, new_folder_path):
        print_on_gui('Criando diretório para os arquivos zip ...')

        # Ensure the original folder exists before attempting to copy
        if not os.path.exists(original_folder_path):
            print_on_gui(f"O diretório original '{original_folder_path}' não existe.")
            return []

        if os.path.isdir(new_folder_path):
            # Use glob to find all ZIP files within the directory
            zip_files_list = glob.glob(os.path.join(new_folder_path, '*.zip'))

            # Check if any ZIP files were found
            if zip_files_list:
                print(f"ZIP files found in directory ({new_folder_path}).")
                return zip_files_list
            else:
                print_on_gui(f"No ZIP files found in the directory ({new_folder_path}).")
                return []

        else:
            print_on_gui(f"Zipando arquivos na pasta '{new_folder_path}'.")
            shutil.copytree(original_folder_path, new_folder_path)
            print_on_gui(f"O diretório '{original_folder_path}' foi duplicado com sucesso como '{new_folder_path}'.")
            zip_files_list = self.send_to_zip(new_folder_path)
            return zip_files_list

    @staticmethod
    def find_client_name_xml(file_path):
        # Check if the file is accessible
        if not os.access(file_path, os.R_OK):
            print_on_gui(f"Error: Permission denied or file does not exist - '{file_path}'")
            return 'Permission Denied or File Not Found'

        try:
            # Parse the XML file
            tree = Et.parse(file_path)
            root = tree.getroot()

            # Namespace map
            namespaces = {'ans': 'http://www.ans.gov.br/padroes/tiss/schemas'}

            # Find the 'ans:nomeBeneficiario' element
            nome_beneficiario_element = root.find('.//ans:nomeBeneficiario', namespaces=namespaces)

            # Check if the element was found and return its text
            if nome_beneficiario_element is not None:
                return nome_beneficiario_element.text
            else:
                return 'VAZIO -'
        except Et.ParseError:
            print_on_gui(f"Error: Could not parse XML - '{file_path}'")
            return 'XML Parse Error'
        except Exception as e:
            print_on_gui(f"An unexpected error occurred: {e}")
            return 'Unexpected Error'

    @staticmethod
    def list_files_in_historico(historico_folder_path):
        file_list = []
        # Check if the 'Historico' folder exists
        if not os.path.exists(historico_folder_path):
            print_on_gui(f"O diretório'{historico_folder_path}' não existe.")
            return file_list

        # Traverse the 'Historico' directory
        for root, dirs, files in os.walk(historico_folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_list.append(file_path)
        return file_list

    @staticmethod
    def get_credentials():
        from dotenv import load_dotenv
        import os

        # Caminho para o seu arquivo .txt
        dotenv_path = r'C:\Users\bi\Documents\BotCase\Scripts\credenciais.txt'

        # Carregar variáveis do arquivo
        load_dotenv(dotenv_path=dotenv_path)

        # Acessar as variáveis
        username = os.getenv('BOT_USERNAME')
        password = os.getenv('BOT_PASSWORD')
        return [username, password]

    @staticmethod
    def send_to_zip_realocating(file_path):
        print_on_gui('Movendo arquivos zip...')

        directory_path = file_path
        created_zip_files = []
        historico_path = os.path.join(directory_path, 'Historico')

        # Get the current month and day names
        current_month = datetime.now().strftime('%B')
        today = datetime.now().strftime('%Y-%m-%d')
        monthly_dir_path = os.path.join(historico_path, current_month)
        today_dir_path = os.path.join(directory_path, f'Data de processamento - {today}')

        # Create 'Historico', monthly and today's directories if they don't exist
        # CHECAR SE O ARQUIVO A SER UTILIZADO EXISTE DE FATO E SEN NÃO, CRIAR UM DIRETORIO
        os.makedirs(monthly_dir_path, exist_ok=True)
        os.makedirs(today_dir_path, exist_ok=True)

        # Iterate over all files in the directory
        for filename in os.listdir(directory_path):
            # client_name = find_client_name_xml(file_path)
            if filename.endswith('.xml'):
                full_file_path = os.path.join(directory_path, filename)

                # Check if it's a file and not a directory
                if os.path.isfile(full_file_path):
                    # Create zip file
                    # new_filename = filename.lstrip('0').split('_')[0]
                    new_filename, nome_beneficiario, numero_lote = get_beneficiario_initials_and_numero_lote(
                        full_file_path)
                    zip_filename = new_filename + '.zip'
                    zip_file_path = os.path.join(today_dir_path, zip_filename)

                    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(full_file_path, os.path.basename(full_file_path))
                    created_zip_files.append(zip_file_path)

                    # Move the original XML file to the monthly directory
                    shutil.move(full_file_path, os.path.join(monthly_dir_path, filename))
                    print_on_gui(f"Arquivo zip criado: {zip_filename} e seu respectivo arquivo XML realocado")

        return created_zip_files

    @staticmethod
    def send_to_zip(directory_path):
        print_on_gui('Zipando arquivos...')

        file_details = []  # List to store the paths of the created ZIP files

        for filename in os.listdir(directory_path):
            if filename.endswith('.xml'):
                full_file_path = os.path.join(directory_path, filename)

                if os.path.isfile(full_file_path):
                    # Here you would call your existing function to get the new filename
                    new_filename, nome_beneficiario, numero_lote = (get_beneficiario_initials_and_numero_lote
                                                                    (full_file_path))
                    zip_filename = new_filename + '.zip'
                    zip_file_path = os.path.join(directory_path, zip_filename)

                    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(full_file_path, os.path.basename(full_file_path))

                    print_on_gui(f"ZIP file created: {zip_filename}")
                    file_details.append([zip_file_path, nome_beneficiario, numero_lote])

        df = pd.DataFrame(file_details, columns=['filepath', 'nome_beneficiario', 'numero_lote'])
        return df

    def close_popup(self, popup_locator, close_popup_button_locator=None, timeout=3):
        print_on_gui('Procurando por popups...')
        try:
            self.su.find_element(popup_locator, timeout=3)
            self.su.click_element(close_popup_button_locator)
            print_on_gui('Popup encontrado e fechado, prosseguindo...')
        except TimeoutException:
            print_on_gui('Popups não encontrados, prosseguindo...')
            pass
        except Exception as e:
            if "GetHandleVerifier" in str(e):
                pass
            print(f'ErrorClosingPopup: {e}')

    def check_warning_messages(self, locator, close_button_locator=None, message=None, timeout=3):
        print_on_gui('Procurando por popups de aviso...')
        warning_element = self.su.find_element(locator=locator,
                                               timeout=timeout)
        if warning_element is not None:
            warning_element = self.su.find_element(locator=locator,
                                                   timeout=timeout)
            if message is not None:
                if not warning_element.text == message:
                    raise Exception(
                        f'OrizonErrorMessage: {message}'
                )
            else:
                self.su.click_element(locator=close_button_locator)
        else:
            print_on_gui('Popups de aviso não encontrados, prosseguindo...')

    def go_to_fature_page(self):
        self.close_popup(popup_locator="//a[contains(@class, 'cc-btn') and contains(text(), 'Permitir todos os Cookies')]",
                         close_popup_button_locator="//a[contains(@class, 'cc-btn') and contains(text(), 'Permitir todos os Cookies')]")

        self.su.click_element('//*[@id="navbarBasicExample"]/div/div[3]/a')
        self.su.click_element("//button[@class='btn-efetuar-login' and @name='Faturi']")

    def faz_login(self, username_str, password_str):
        # XPath elements
        print_on_gui('Fazendo login no portal Orizon...')

        username_field = self.su.find_element('username', 'id')
        password_field = self.su.find_element("password", 'id')
        acessar_button = self.su.find_element('kc-login', 'id')

        username_field.send_keys(username_str)
        password_field.send_keys(password_str)
        acessar_button.click()

        self.check_warning_messages(locator='//span[text() = "Nome de usuário ou senha inválidos."]',
                                    message='Nome de usuário ou senha inválidos.')

    def upload_xml(self, zip_files_path_dataframe, output_dataframe, output_filepath):
        print_on_gui('Iniciando envio de arquivos...')

        time.sleep(1)
        self.close_popup(popup_locator="//button[text()='Terminar']",
                             close_popup_button_locator="//button[text()='Terminar']")

        self.su.click_element("//span[contains(text(), 'Enviar XML TISS')]")
        time.sleep(1)

        zip_files_path_list = zip_files_path_dataframe['filepath']
        for file_path in zip_files_path_list:

            nome_beneficiario = self.filter_df(zip_files_path_dataframe, 'filepath', file_path)['nome_beneficiario'].to_string(
                index=False)
            numero_lote = self.filter_df(zip_files_path_dataframe, 'filepath', file_path)['numero_lote'].to_string(index=False)
            filename = os.path.basename(file_path).lstrip('0').split('_')[0]

            def makes_the_upload(output_dataframe_upload, file_name, filepath):
                try:
                    # Wait for any overlay to disappear
                    #self.su.find_element("div.overlay", 'css')

                    # Then wait for the actual 'input' element to be present
                    # It might be hidden, so we don't wait for it to be clickable, just present
                    self.su.find_element("//input[@type='file']")

                    file_input = self.su.find_element("input[type='file']", 'css')
                    file_input.send_keys(file_path)

                    time.sleep(5)
                    # Update your output dataframe
                    info = {
                        'Numero_lote': str(numero_lote),
                        'Nome_beneficiario': nome_beneficiario,
                        'File Path': filepath,
                        'File Name': file_name,
                        'Upload Date': datetime.now().strftime('%d/%m/%Y %H:%M'),
                        'Status': 'Uploaded'
                    }
                    # Convert info dictionary to DataFrame
                    info_df = pd.DataFrame([info])
                    # Use pd.concat to append the new row
                    output_dataframe_upload = pd.concat([output_dataframe_upload, info_df], ignore_index=True)
                    return output_dataframe_upload, info
                except Exception as error:
                    print(f"Error while trying to upload files: {error}")
                    return output_dataframe, None

            output_dataframe, data = makes_the_upload(output_dataframe, filename, file_path)

            if self.su.find_element('//td/small'):
                message = self.su.find_element('//td/small').text
                data['Status'] = f"Error: {message}"
                print_on_gui(f'Problema ao enviar arquivo {filename}. \n Erro: {message}')
                self.su.driver.refresh()
                data_df = pd.DataFrame([data])
                output_dataframe = pd.concat([output_dataframe, data_df], ignore_index=True)
                time.sleep(5)
                continue
            self.su.click_element("//button[@class='btn btn-default orange btn-inner' and @ng-click='openModalConfirmaTiss()' and @id='botaoAssinarEnviar']")

            try:
                time.sleep(1)
                # Colocar if clickable then voltar depois
                if not self.su.find_element(locator='//span[contains(text(), "Arquivo válido para envio.")]'):
                    self.su.click_element(locator='//button[contains(@class, "btn-default") and contains(text(), "Exportar erros dos arquivos")]')
                    data = {
                        'File Path': file_path,
                        'File Name': filename,
                        'Upload Date': datetime.now().strftime('%d/%m/%Y %H:%M'),
                        'Status': 'Erro de agendamento'
                    }

                    # Convert the dictionary to a DataFrame
                    data_df = pd.DataFrame([data])

                    # Use pd.concat to append the new row
                    output_dataframe = pd.concat([output_dataframe, data_df], ignore_index=True)

                    self.su.click_element('//*[@id="botaoConfirmarEnvio"]')
                    print_on_gui(f'Tentativa de envio do arquivo {filename} falhou por erro de agendamento.')
                    continue
            except (TimeoutException, NoSuchElementException):
                print_on_gui('Envio do arquivo feito com sucesso.')


            self.su.click_element("//button[contains(@class, 'btn-default') and contains(text(), 'Enviar arquivos válidos')]")

            # Tentar clicar duas vezes pois o site pode travar ao clicar apenas uma
            self.su.click_element('//*[@id="botaoEnviarMais"]')
            try:
                self.su.click_element('//*[@id="botaoEnviarMais"]', timeout=5)
            except TimeoutException:
                pass

            #TODO: FIX THIS
            # self.append_data_to_excel(output_filepath, data)

            time.sleep(3)

            print_on_gui(f'Remessa {filename} enviada.')

        return output_dataframe

    def select_all_to_send_and_send(self, aba='Enviar'):
        time.sleep(3)
        try:
            self.su.wait_for_element('//*[@id="paginationBox_item1"]/small')
            number_of_lots = self.su.find_element('//*[@id="paginationBox_item1"]/small').text
            if number_of_lots is None:
                print_on_gui('Nenhum dos lotes foi enviado ou o portal Orizon não os carregou corretamente.')
                self.su.driver.refresh()
                return
            number_of_lots_displayed = self.su.find_element('//*[@id="paginationBox_item1"]/div/input')
            number_of_lots_displayed.clear()
            number_of_lots_displayed.send_keys(self.extract_last_number(number_of_lots))
            self.su.click_element('/html/body/div[1]')

            # Wait for the 'Select All' checkbox to become clickable and send it
            time.sleep(5)
            self.su.click_element('//th/input', simple=True)
            self.su.click_element(f'//*[@id="botao{aba}LotesSelecionados"]')

        except Exception as e:
            if 'InvalidLots' in str(e):
                print_on_gui('Os lotes enviados não foram aceitos...encerrando o programa.')
                return False
            print_on_gui(f'An error occurred while sending masses: \n {e}')

    def send_masses(self):
        print_on_gui('Confirmando envio dos arquivos...')

        self.close_popup(popup_locator="//button[text()='Terminar']",
                         close_popup_button_locator="//button[text()='Terminar']")

        # Envia lotes novos
        self.su.click_element('//a[contains(text(), "Lista de Lotes")]')
        self.select_all_to_send_and_send('Enviar')
        self.su.click_element('//div[2]/button[contains(text(), "Ir para lotes recentes") and @id="botaoEnviarLotesSelecGotoLotesRecentes"]')

        # Vai para a aba de Recebidos e libera os lotes
        self.su.click_element('//span[contains(text(), "Recebidos")]')
        self.select_all_to_send_and_send('Liberar')
        self.su.click_element('//div[3]/button[contains(text(), "Confirmar liberação")]')

        # Vai para a aba de liberação e clica para liberar os lotes
        self.su.click_element('//span[contains(text(), "Liberados")]')



    def main(self, file_path):
        output_filepath = None
        self.su = SeleniumUtils(driver_path=self.driver_path)
        self.su.open_url(self.orizon_url)
        if file_path == '':
            raise ValueError('Favor informar o caminho da pasta com os arquivos XML a serem enviados. \n')

        try:
            # Example usage:
            original_folder = r'C:\Users\bi\Documents\BotCase\Dash\Teste Dir\teste template'
            # if prod:
            if not os.path.isdir(file_path):
                shutil.copytree(original_folder, file_path)
                zip_files_path_dataframe = self.send_to_zip(file_path)
            else:
                zip_files_path_dataframe = self.send_to_zip(file_path)

            self.faz_login(self.username, self.password)

            self.go_to_fature_page()

            self.close_popup(popup_locator='//aside/div/div/div/div/div/img',
                             close_popup_button_locator='//div[@class="modal-buttons inline"]//button[contains(text(), "Fechar")]',
                             timeout=7)

            output_dataframe = pd.DataFrame(['Client Name', 'Upload Date', 'File Name', 'File Path'])

            output_filepath = self.generate_filename()
            output_dataframe = self.upload_xml(zip_files_path_dataframe, output_dataframe, output_filepath)

            writer = pd.ExcelWriter(r'C:\Users\bi\Downloads\Output.xlsx', engine='openpyxl')
            output_dataframe.to_excel(writer, sheet_name='Output_Dataframe')

            self.send_masses()

            df = pu.get_or_create_excel(output_filepath)
            df = df.drop_duplicates(keep='last')

            process_files(directory_to_organize=file_path, df=df)
            print_on_gui('Arquivos enviados com sucesso!')
            time.sleep(10 * 60 * 60)



        except urllib3.exceptions.ProtocolError as e:
            e = str(e)
            # Check if the specific error message is in the exception
            if 'Connection aborted.' in e or 'ConnectionResetError(10054' in e or 'NoSuchWindowException' in e:
                print_on_gui('A janela principal foi fechada. Por favor selecione os arquivos novamente e tente de novo.')

        except Exception as e:
            if output_filepath is not None:
                df = pu.get_or_create_excel(output_filepath)
                df = df.drop_duplicates(keep='last')
                process_files(directory_to_organize=file_path, df=df)
            raise Exception(f'ErroronMainFuncion: {e}')

        #TODO: Understand where the df is created and choose where would it be better for the bot to write to the dataframe
